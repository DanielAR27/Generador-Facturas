"""
=============================================================================
Generador de Facturas de Evaluación (Rúbricas Automatizadas)
=============================================================================

Desarrollado por: Daniel Alemán Ruiz
Institución: Instituto Tecnológico de Costa Rica (TEC)
Uso: Exclusivo para fines académicos y asistencia docente.

Descripción: 
Aplicación web desarrollada con Streamlit para automatizar la creación 
de rúbricas en formato Excel. Permite la lectura de listas de estudiantes, 
creación dinámica de equipos, configuración de parámetros de evaluación 
(partes y pesos), y la inyección de datos y fórmulas en una plantilla base.
=============================================================================
"""

import streamlit as st
import pandas as pd
import os
import openpyxl
from copy import copy
import re

# Configuración de la página
st.set_page_config(page_title="Generador de Facturas", page_icon="📝", layout="centered")

st.title("Generador de Facturas de Evaluación")

@st.cache_data
def cargar_cursos(ruta_archivo):
    if os.path.exists(ruta_archivo):
        return pd.read_csv(ruta_archivo, sep=';', quotechar='"')
    else:
        return pd.DataFrame()

def copiar_formato(celda_origen, celda_destino):
    if celda_origen.has_style:
        celda_destino.font = copy(celda_origen.font)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.alignment = copy(celda_origen.alignment)

def limpiar_nombre_archivo(nombre):
    # Elimina los caracteres que Windows odia 
    # y cambia espacios por guiones bajos
    nombre_limpio = re.sub(r'[<>:"/\\|?*]', '', nombre)
    return nombre_limpio.strip().replace(" ", "_")

st.subheader("1. Selección de Curso")
df_cursos = cargar_cursos("cursos.csv")

if not df_cursos.empty:
    opciones_cursos = df_cursos['Siglas'] + " - " + df_cursos['Curso']
    curso_seleccionado = st.selectbox("Selecciona el curso a evaluar:", opciones_cursos)
    siglas_curso = curso_seleccionado.split(" - ")[0] 
    
    st.subheader("2. Selección de Estudiantes")
    
    archivos_locales = os.listdir('.')
    archivos_estudiantes = [f for f in archivos_locales if f.endswith('.csv') and f != 'cursos.csv']
    
    if archivos_estudiantes:
        archivo_seleccionado = st.selectbox("Selecciona el archivo de estudiantes:", archivos_estudiantes)
        
        try:
            df_estudiantes = pd.read_csv(archivo_seleccionado, sep=';', quotechar='"')
            
            if 'Rol' in df_estudiantes.columns:
                df_estudiantes = df_estudiantes[df_estudiantes['Rol'] == 'student']
            
            if 'Apellidos' in df_estudiantes.columns and 'Nombre' in df_estudiantes.columns:
                df_estudiantes['Nombre Completo'] = df_estudiantes['Apellidos'] + " " + df_estudiantes['Nombre']
                lista_estudiantes = df_estudiantes['Nombre Completo'].tolist()
                
                col1, col2 = st.columns(2)
                with col1:
                    modo_evaluacion = st.radio("Modo de Evaluación:", ["Individual", "Por Equipos"])
                
                grupos_a_iterar = [] 

                if modo_evaluacion == "Individual":
                    with col2:
                        st.write("\n")
                        seleccionar_todos = st.checkbox("Seleccionar a todos los estudiantes")

                    if seleccionar_todos:
                        estudiantes_seleccionados = st.multiselect("Estudiantes a evaluar:", lista_estudiantes, default=lista_estudiantes)
                    else:
                        estudiantes_seleccionados = st.multiselect("Estudiantes a evaluar:", lista_estudiantes)
                    
                    if estudiantes_seleccionados:
                        grupos_a_iterar = [[est] for est in estudiantes_seleccionados]

                elif modo_evaluacion == "Por Equipos":
                    st.write("---")
                    num_equipos = st.number_input("Cantidad de equipos a formar:", min_value=1, max_value=20, value=1)
                    
                    st.write("Selecciona los integrantes de cada equipo:")
                    
                    # 1. Recopilar todos los estudiantes que están asignados en la memoria
                    estudiantes_en_uso = set()
                    for i in range(num_equipos):
                        key = f"equipo_{i}"
                        if key in st.session_state:
                            estudiantes_en_uso.update(st.session_state[key])

                    # 2. Se generan las cajas filtrando las opciones
                    for i in range(num_equipos):
                        key = f"equipo_{i}"
                        seleccion_actual = st.session_state.get(key, [])
                        
                        # Los disponibles son: Todos - Los que están en uso + Los que ya metí en este equipo específico
                        opciones_disponibles = [
                            est for est in lista_estudiantes 
                            if est not in estudiantes_en_uso or est in seleccion_actual
                        ]
                        
                        equipo = st.multiselect(f"Equipo {i+1}:", options=opciones_disponibles, key=key)
                        if equipo:
                            grupos_a_iterar.append(equipo)

                if grupos_a_iterar:
                    st.divider()
                    st.subheader("3. Configuración de la Tarea")
                    
                    nombre_tarea = st.text_input("Nombre de la tarea:", "Tarea 1")
                    
                    # NUEVO: Etiqueta personalizada
                    etiqueta_parte = st.text_input("¿Cómo llamar a las divisiones? (Ej: Parte, Ejercicio, Reto):", "Parte")
                    
                    # --- Lógica de autocompletado de pesos ---
                    if "prev_num_partes" not in st.session_state:
                        st.session_state.prev_num_partes = 1

                    num_partes = st.number_input(f"Cantidad de {etiqueta_parte.lower()}s:", min_value=1, max_value=20, value=1, key="num_partes_input")

                    # Si el usuario cambia el número de partes, se reparte el 100% equitativamente
                    if num_partes != st.session_state.prev_num_partes:
                        peso_eq = 100.0 / num_partes
                        for i in range(num_partes):
                            st.session_state[f"peso_{i}"] = peso_eq
                        st.session_state.prev_num_partes = num_partes
                    # ------------------------------------------------------------------
                    
                    pesos = []
                    nombres_especificos = []
                    
                    st.write(f"### Configuración de {etiqueta_parte}s")
                    
                    with st.expander(f"Ver detalles de cada {etiqueta_parte.lower()}", expanded=True):
                        if num_partes == 1:
                            if "peso_0" not in st.session_state:
                                st.session_state["peso_0"] = 100.0
                            peso = st.number_input("Peso (%)", min_value=0.0, max_value=100.0, step=0.5, key="peso_0")
                            nombre_esp = st.text_input(f"Nombre del {etiqueta_parte} 1:", placeholder="Ej: Algoritmos de búsqueda")
                            pesos.append(peso)
                            nombres_especificos.append(nombre_esp)
                        else:
                            for i in range(num_partes):
                                st.markdown(f"**{etiqueta_parte} {i+1}**")
                                if f"peso_{i}" not in st.session_state:
                                    st.session_state[f"peso_{i}"] = 100.0 / num_partes
                                
                                peso = st.number_input("Peso (%)", min_value=0.0, max_value=100.0, step=0.5, key=f"peso_{i}")
                                nombre_esp = st.text_input("Nombre", key=f"nom_{i}", placeholder=f"Opcional para {etiqueta_parte.lower()} {i+1}")
                                
                                pesos.append(peso)
                                nombres_especificos.append(nombre_esp)
                                
                                if i < num_partes - 1:
                                    st.divider()
                    
                    if abs(sum(pesos) - 100.0) > 0.1:
                        st.warning(f"⚠️ La suma total de los porcentajes es {sum(pesos):.1f}%. Lo ideal es que sea 100%.")

                    carpeta_salida = st.text_input("Carpeta para guardar los archivos:", "Resultados_Facturas")

                    if st.button(" Generar Facturas", type="primary"):
                        if not os.path.exists("plantilla.xlsx"):
                            st.error("❌ No se encontró el archivo 'plantilla.xlsx'.")
                        else:
                            if not os.path.exists(carpeta_salida):
                                os.makedirs(carpeta_salida)

                            try:
                                barra_progreso = st.progress(0)
                                
                                for idx, grupo in enumerate(grupos_a_iterar):
                                    wb = openpyxl.load_workbook("plantilla.xlsx")
                                    ws = wb.active

                                    # 1. Modificar encabezados básicos
                                    ws['B2'] = curso_seleccionado
                                    ws['B3'] = nombre_tarea
                                    
                                    if modo_evaluacion == "Por Equipos":
                                        ws['A5'] = "Estudiantes:"
                                        ws['B5'] = ", ".join(grupo)
                                        apellidos = [est.split(" ")[0] for est in grupo]
                                        nombre_estudiantes_archivo = "Equipo_" + "_".join(apellidos)
                                    else:
                                        ws['A5'] = "Estudiante:"
                                        ws['B5'] = grupo[0]
                                        nombre_estudiantes_archivo = grupo[0].replace(" ", "_")

                                    # 2. Modificar partes (Fila 9 en adelante)
                                    alineacion_neta = copy(ws['D11'].alignment)
                                    
                                    if num_partes > 1:
                                        # Descombinar celdas originales antes de insertar
                                        for row_idx in [10, 11, 12]:
                                            try:
                                                ws.unmerge_cells(f'D{row_idx}:H{row_idx}')
                                            except Exception:
                                                pass
                                                
                                        ws.insert_rows(10, amount=num_partes - 1)
                                        
                                        # Limpiador de celdas fantasma (Destruye uniones accidentales como E13:H13)
                                        fusiones_a_borrar = []
                                        for merge in list(ws.merged_cells.ranges):
                                            if 9 <= merge.min_row <= 9 + num_partes:
                                                fusiones_a_borrar.append(str(merge))
                                        
                                        for fusion in fusiones_a_borrar:
                                            try:
                                                ws.unmerge_cells(fusion)
                                            except Exception:
                                                pass
                                    
                                    for i in range(num_partes):
                                        fila_actual = 9 + i
                                        texto_celda = f"{etiqueta_parte} {i+1}"
                                        if nombres_especificos[i].strip():
                                            texto_celda += f": {nombres_especificos[i]}"
                                            
                                        ws[f'A{fila_actual}'] = texto_celda
                                        ws[f'B{fila_actual}'] = pesos[i]
                                        
                                        ws[f'J{fila_actual}'] = f"=(C{fila_actual}*$C$8+D{fila_actual}*$D$8+E{fila_actual}*$E$8+F{fila_actual}*$F$8+G{fila_actual}*$G$8+H{fila_actual}*$H$8)*B{fila_actual}/100"
                                        
                                        if i > 0:
                                            for col in range(1, 12): 
                                                celda_origen = ws.cell(row=9, column=col)
                                                celda_destino = ws.cell(row=fila_actual, column=col)
                                                copiar_formato(celda_origen, celda_destino)

                                    # 3. Actualizar fórmulas desplazadas y forzar Merge
                                    fila_nota_neta = 9 + num_partes + 1
                                    fila_nota_final = 9 + num_partes + 3
                                    fila_pesos_tardia = 9 + num_partes + 4
                                    fila_checks_tardia = 9 + num_partes + 6
                                    fila_suma_tardia = 9 + num_partes + 8

                                    try:
                                        ws.merge_cells(f'D{fila_nota_neta}:H{fila_nota_neta}')
                                        ws[f'D{fila_nota_neta}'].alignment = alineacion_neta 
                                    except Exception:
                                        pass

                                    ws[f'J{fila_nota_neta}'] = f"=SUM(J9:J{fila_nota_neta-1})"
                                    ws[f'F{fila_suma_tardia}'] = f"=D{fila_checks_tardia}*D{fila_pesos_tardia}+E{fila_checks_tardia}*E{fila_pesos_tardia}+F{fila_checks_tardia}*F{fila_pesos_tardia}+G{fila_checks_tardia}*G{fila_pesos_tardia}+H{fila_checks_tardia}*H{fila_pesos_tardia}"
                                    ws[f'J{fila_nota_final}'] = f"=J{fila_nota_neta}-J{fila_nota_neta}*F{fila_suma_tardia}"

                                    # 4. Guardar archivo
                                    nombre_tarea_limpio = limpiar_nombre_archivo(nombre_tarea)
                                    nombre_archivo = f"Factura_{nombre_tarea_limpio}_{siglas_curso}_{nombre_estudiantes_archivo}.xlsx"
                                    ruta_guardado = os.path.join(carpeta_salida, nombre_archivo)
                                    
                                    wb.save(ruta_guardado)
                                    barra_progreso.progress((idx + 1) / len(grupos_a_iterar))

                                st.success(f"✅ ¡Se generaron {len(grupos_a_iterar)} archivo(s) con éxito en la carpeta '{carpeta_salida}'!")
                                
                            except Exception as e:
                                st.error(f"Error al generar los archivos: {e}")

            else:
                st.error("El archivo no tiene las columnas 'Apellidos' y 'Nombre'.")
        except Exception as e:
            st.error(f"Error al procesar el archivo: {e}")